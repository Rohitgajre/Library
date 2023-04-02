from django.shortcuts import render, HttpResponse, redirect
from .models import Book
# Create your views here.
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

@csrf_exempt
@login_required
def home(request):
    print(request.POST)
    if request.method =="POST":
        data = request.POST.get
        print(data)
        bid = data("book_id")
        name = data("book_name")
        qty = data("book_qty")
        price = data("book_price")
        author = data("book_author") # all the field  Getting data from the client i.e browser
        # print(request.POST)
        is_pub = request.POST.get("book_is_pub")
        # print(name, qty, price, author, is_pub)
        if is_pub == "Yes":
            is_pub = True
        else:
            is_pub = False 
        if not bid:
            Book.objects.create(name=name, qty=qty, price=price, author=author, is_published=is_pub) # To storing the data into the database using ORM queary it will run on backend
        else:
            book_obj = Book.objects.get(id=bid)
            book_obj.name = name
            book_obj.qty = qty
            book_obj.price = price
            book_obj.author = author
            book_obj.is_published = is_pub
            book_obj.save()
        # return HttpResponse("Success")
        return redirect("home_page")
    elif request.method == "GET":
        # print(request.GET)
        # return render(request, "old_home.html", context={"person_name": "Rohit "}) # Dynamica data which we can pass through html to clinet or on browoser
        # return render(request, "home.html", context={"person_name":["ABC", "xyz", "pqr"]}) # Dynamica data which we can pass through html to clinet or on browoser
        return render(request, "old_home.html", context={"all_books": Book.objects.all()} ) # Dynamica data which we can pass through html to clinet or on browoser

@login_required
def show_books(request):
    return render(request, "show_book.html", {"books": Book.objects.filter(is_active=True), "active": True})

@login_required
def update_book(request, pk):        # second parameter in manditory with request
    book_obj = Book.objects.get(id=pk)
    return render(request, "home.html", context={"single_book": book_obj})

@login_required
def delete_book(request, pk):
    Book.objects.get(id=pk).delete()
    return redirect("all_active_books")

@login_required
def soft_delete_book(request, pk):
    book_obj = Book.objects.get(id=pk)
    book_obj.is_active = False
    book_obj.save()
    return redirect("all_inactive_books")

@login_required
def show_inactive_books(request):
    # return render(request, "show_books.html", {"books": Book.objects.filter(is_active=False)})
    return render(request, "show_book.html", {"books": Book.objects.filter(is_active=False), "inactive": True})

@login_required
def restore_book(restore, pk):
    book_obj = Book.objects.get(id=pk)
    book_obj.is_active = True
    book_obj.save()
    return redirect("all_active_books")


from .forms import BookForm, AddressForm
from django.contrib.auth.forms import UserCreationForm

@login_required
def book_form(request):
    form = BookForm()
    if request.method == "POST":
        print(request.POST)
        form = BookForm(data = request.POST)
        if form.is_valid():
            print(form.cleaned_data)
            form.save()
            return HttpResponse("Sucessfully")
    else:
        context = {'form': form}
        return render(request, "book_form.html", context=context)

@login_required
def sibtc(request):
    return render(request, 'sibtc.html', {"form": AddressForm()})


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def index(request):
    book_list = Book.objects.all()
    page = request.GET.get('page', 1)
    paginator = Paginator(book_list, 3)
    try:
        books = paginator.page(page)
    except PageNotAnInteger:
        books = paginator.page(1)
    except EmptyPage:
        books = paginator.page(paginator.num_pages)
    return render(request, 'index.html', {'books': books})


# from django.views import View 

# class NewView(View):  
#     def get(self, request):  
#         # View logic will place here  
#         return HttpResponse('get response')  
    
#     def post(self, request):
#         return HttpResponse("post response")

#     def put(self, request): # update
#         return HttpResponse("put response")

#     def patch(self, request): # partial info update
#         return HttpResponse("patch response")

#     def delete(self, request): # delete
#         return HttpResponse("delete response")





# # CRUD
from django.views.generic.edit import CreateView
from django.urls import reverse, reverse_lazy  

class BookCreate(CreateView):  # get/post handled
    model = Book  
    fields = '__all__'  
    # success_url = "/cbv-create-book/"  #  
    success_url = reverse_lazy('BookCreate') 

from django.views.generic.list import ListView  
  
class BookRetrieve(ListView):  
    model = Book 
    context_object_name = "all_books"
    # http_method_names = ['get', 'post', 'put', 'patch', 'delete', 'head', 'options', 'trace']
    queryset = Book.objects.filter(is_active=1)

    def get_queryset(self):
        print("in method")
        return Book.objects.filter(is_active=0)

from django.views.generic.detail import DetailView  
  
class BookDetail(DetailView):  # here need provide id in after url 
    model = Book  


from django.views.generic.edit import UpdateView, DeleteView  

class BookUpdate(UpdateView):  
    model = Book 
    fields = "__all__"
    success_url = "/cbv-create-book/"


class BookDelete(DeleteView):
    model = Book
    success_url = "/cbv-create-book/"



from django.views.generic import TemplateView

class Template(TemplateView):
    template_name  = "home.html"  # {{name}}
    extra_context = {"name": "Aakash"}



from django.http import HttpResponse
import csv

def create_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="test.csv"'

    writer = csv.writer(response)
    writer.writerow(['name','qty', 'price', 'author', 'is_published', 'is_active'])

    books = Book.objects.all().values_list('name','qty', 'price', 'author', 'is_published', 'is_active')
    for book in books:
        writer.writerow(book)
    return response



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib.auth.models import User

def export_xls(request):
    books = Book.objects.all()
    wb = Workbook()

    ws_active = wb.active
    ws_active.title = "Active Data"
    headers = ["Name", "qty", "price", "author", "is_published"]
    ws_active.append(headers)
    ws_inactive = wb.create_sheet(title="InActive Data")
    ws_inactive.append(headers)
    for book in books:
        row_data = [book.name, book.qty, book.price,book.author, book.is_published]
        if book.is_published:
            ws_active.append(row_data)
        else:
            ws_inactive.append(row_data)

    response = HttpResponse(content_type = "application/vnd.ms-excel")
    response["content-Dispostion"] = "attachment; filename = books.xlsx"
    wb.save(response)
    return response
    


    

# # Assignement:- 9th   -- 
# # book csv export
# # excel -- Active books shaeet- active books, inactive sheet-inactive books, 
# # raw queries - using objects.raw  (select * from books;) -- csv me dalna
# # read text file and show its content on UI using view
# # download sample csv file
# # validations - duplicate book not allowed


def upload_csv(request): 
    file = request.FILES["csv_file"]         
    decoded_file = file.read().decode('utf-8').splitlines()
    expected_header_lst = ['name', "qty", "price", "author", "is_published"]
    expected_header_lst.sort()

    actual_header_lst = decoded_file[0].split(",")
    actual_header_lst.sort()
    # print(expected_header_lst, actual_header_lst)
    if expected_header_lst != actual_header_lst:
        return HttpResponse("Error...Headers are not equal..!")

    reader = csv.DictReader(decoded_file) # always use DictReader
    lst = []
    for element in reader:
        print(element)
        is_pub = element.get("is_published")
        if is_pub == "TRUE":
            is_pub = True
        else:
            is_pub = False
        lst.append(Book(name=element.get("name"), qty=element.get("qty"), price=element.get("price"), author=element.get("author"), is_published=is_pub))
    print(reader)
    Book.objects.bulk_create(lst)
    return HttpResponse("Success")
